"""
 Command for mailing all the data
"""
from django.core.management.base import BaseCommand
from main.models import *

import openpyxl 


class Command(BaseCommand):
    help = " Command for mailing all the data"
    requires_system_checks = output_transaction = True

    def add_arguments(self, parser):
        # Positional arguments
        parser.add_argument("game_id", nargs="+", type=str)

    def handle(self, *args, **options):
        if options.get('game_id') is not None:
            game_obj = Games.objects.filter(game_unique_id=options.get('game_id')[0]).get()
            
            wb = openpyxl.Workbook() 
            sheet = wb.active 
            
            c1 = sheet.cell(row = 1, column = 1) 
            c1.value = "Group Name"
            
            c2 = sheet.cell(row = 1, column = 2) 
            c2.value = "Group Unique ID"
            
            c3 = sheet.cell(row = 1, column = 3) 
            c3.value = "Mode"
            
            c4 = sheet.cell(row = 1, column = 4) 
            c4.value = "Game Name"
            
            c5 = sheet.cell(row = 1, column = 5) 
            c5.value = "Payment Id"
            
            c6 = sheet.cell(row = 1, column = 6) 
            c6.value = "Group Members Details"

            data = []
            for game_group in GameGroup.objects.filter(game__id=game_obj.id).iterator():
                user_data = ''
                user_format = 'Name: {firstname} {lastname}\n Address: {address}\n Phone No: {phone_no}\n Email Address: {email}\n University: {university} \n Referral Code: {referral_code} \n Gender: {gender} \n Unique ID: {unique_id}\n\n'
                for i in game_group.users.all():
                    user_data_format = str(user_format)
                    user_data+=user_data_format.format(
                        firstname=i.first_name, 
                        lastname=i.last_name, 
                        address=str(i.address1)+'\n'+str(i.address2)+'\n'+str(i.city)+'\n'+str(i.state)+'\n'+str(i.country),
                        phone_no=i.phone,
                        email=i.email,
                        university=i.university_name,
                        referral_code=i.referral_code,
                        gender=i.gender,
                        unique_id=str(i.unique_id)
                    )
                    data.append((
                        game_group.group_name,
                        str(game_group.group_unique_id),
                        'SQUAD' if game_group.solo_or_squad == 'sq' else 'SOLO',
                        game_group.game.name,
                        str(game_group.payment_id.order_id),
                        user_data
                    ))
            for row in data:
                sheet.append(row)
            wb.save(f"{game_obj.name}.xlsx") 
        else:
            raise RunTimeError('Please provide the game_id')

